"""
Servicio de exportación de datos de salud.

Genera reportes en PDF, CSV y Excel con los registros
y estadísticas del usuario.
"""

from __future__ import annotations

import csv
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from database.connection import obtener_gestor
from database.repositories.registro_repository import RepositorioRegistro
from core.config import config
from core.constants import APP_NOMBRE, APP_VERSION, PERIODOS
from core.exceptions import ExportacionError

logger = logging.getLogger("healthtrack.services.exportacion")


class ServicioExportacion:
    """
    Genera exportaciones en múltiples formatos.

    Todos los archivos se guardan en el directorio de exportaciones
    configurado por el usuario.
    """

    def __init__(self) -> None:
        self._gestor = obtener_gestor()
        self._directorio = config.exportacion_directorio
        self._directorio.mkdir(parents=True, exist_ok=True)

    # ──────────────────────────────────────────
    # CSV
    # ──────────────────────────────────────────

    def exportar_csv(
        self,
        fecha_inicio: Optional[date] = None,
        fecha_fin: Optional[date] = None,
        nombre_archivo: Optional[str] = None,
    ) -> Path:
        """
        Exporta registros a CSV.

        Args:
            fecha_inicio: Inicio del rango (None = desde el primer registro).
            fecha_fin: Fin del rango (None = hasta hoy).
            nombre_archivo: Nombre personalizado del archivo (sin extensión).

        Returns:
            Ruta al archivo CSV generado.
        """
        try:
            registros = self._obtener_registros(fecha_inicio, fecha_fin)

            nombre = nombre_archivo or self._generar_nombre("registros", "csv")
            ruta = self._directorio / nombre

            encabezados = [
                "ID", "Fecha", "Período",
                "Presión Sistólica (mmHg)", "Presión Diastólica (mmHg)",
                "Ritmo Cardíaco (bpm)", "Oxigenación SpO2 (%)",
                "Peso (kg)", "Altura (cm)", "IMC",
                "Pasos", "Distancia Caminada (km)", "Calorías Quemadas",
                "Horas de Sueño", "Calidad del Sueño (1-10)",
                "Nivel de Estrés (1-10)", "Estado de Ánimo (1-10)",
                "Glucosa (mg/dL)", "Temperatura (°C)",
                "Consumo de Agua (L)", "Cafeína (mg)",
                "Medicamentos", "Síntomas", "Notas Médicas",
                "Ejercicio Realizado", "Criticidad General",
            ]

            with ruta.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(encabezados)

                for r in registros:
                    writer.writerow([
                        r.id,
                        r.fecha.isoformat() if r.fecha else "",
                        PERIODOS.get(r.periodo, r.periodo),
                        r.presion_sistolica or "",
                        r.presion_diastolica or "",
                        r.ritmo_cardiaco or "",
                        r.oxigenacion or "",
                        r.peso or "",
                        r.altura or "",
                        r.imc or "",
                        r.pasos or "",
                        r.distancia_caminada or "",
                        r.calorias_quemadas or "",
                        r.horas_sueno or "",
                        r.calidad_sueno or "",
                        r.nivel_estres or "",
                        r.estado_animo or "",
                        r.glucosa or "",
                        r.temperatura_corporal or "",
                        r.consumo_agua or "",
                        r.cafeina or "",
                        ", ".join(r.medicamentos_lista),
                        ", ".join(r.sintomas_lista),
                        r.notas_medicas or "",
                        r.ejercicio_realizado or "",
                        r.criticidad_general(),
                    ])

            logger.info("CSV exportado: %s (%d registros)", ruta, len(registros))
            return ruta

        except (OSError, Exception) as e:
            raise ExportacionError("CSV", str(e)) from e

    # ──────────────────────────────────────────
    # Excel (XLSX)
    # ──────────────────────────────────────────

    def exportar_excel(
        self,
        fecha_inicio: Optional[date] = None,
        fecha_fin: Optional[date] = None,
        nombre_archivo: Optional[str] = None,
    ) -> Path:
        """Exporta registros a Excel con formato y colores."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

        except ImportError as e:
            raise ExportacionError("Excel", "openpyxl no está instalado") from e

        try:
            registros = self._obtener_registros(fecha_inicio, fecha_fin)

            nombre = nombre_archivo or self._generar_nombre("registros", "xlsx")
            ruta = self._directorio / nombre

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Registros de Salud"

            # Estilos
            color_header = "1e293b"
            color_normal = "22c55e"
            color_atencion = "f59e0b"
            color_critico = "ef4444"

            fuente_header = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            relleno_header = PatternFill("solid", fgColor=color_header)
            alineacion_centro = Alignment(horizontal="center", vertical="center")

            encabezados = [
                "Fecha", "Período", "P. Sistólica", "P. Diastólica",
                "Ritmo Cardíaco", "SpO2 %", "Peso kg", "IMC",
                "Pasos", "Sueño h", "Estrés", "Ánimo",
                "Glucosa", "Temperatura", "Agua L", "Criticidad",
            ]

            # Encabezados
            for col, titulo in enumerate(encabezados, 1):
                celda = ws.cell(row=1, column=col, value=titulo)
                celda.font = fuente_header
                celda.fill = relleno_header
                celda.alignment = alineacion_centro

            # Datos
            colores_criticidad = {
                "normal": "d4edda",
                "atencion": "fff3cd",
                "preocupante": "ffe0cc",
                "critico": "f8d7da",
            }

            for fila_idx, r in enumerate(registros, 2):
                datos = [
                    r.fecha.isoformat() if r.fecha else "",
                    PERIODOS.get(r.periodo, r.periodo),
                    r.presion_sistolica,
                    r.presion_diastolica,
                    r.ritmo_cardiaco,
                    r.oxigenacion,
                    r.peso,
                    r.imc,
                    r.pasos,
                    r.horas_sueno,
                    r.nivel_estres,
                    r.estado_animo,
                    r.glucosa,
                    r.temperatura_corporal,
                    r.consumo_agua,
                    r.criticidad_general().capitalize(),
                ]

                criticidad = r.criticidad_general()
                color_fila = colores_criticidad.get(criticidad, "FFFFFF")
                relleno_fila = PatternFill("solid", fgColor=color_fila)

                for col_idx, valor in enumerate(datos, 1):
                    celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
                    celda.fill = relleno_fila
                    celda.alignment = alineacion_centro

            # Ajuste de columnas
            for col in range(1, len(encabezados) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 14

            wb.save(ruta)
            logger.info("Excel exportado: %s (%d registros)", ruta, len(registros))
            return ruta

        except Exception as e:
            raise ExportacionError("Excel", str(e)) from e

    # ──────────────────────────────────────────
    # PDF
    # ──────────────────────────────────────────

    def exportar_pdf(
        self,
        fecha_inicio: Optional[date] = None,
        fecha_fin: Optional[date] = None,
        nombre_archivo: Optional[str] = None,
    ) -> Path:
        """Exporta un reporte PDF con resumen y tabla de registros."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer, HRFlowable,
            )
        except ImportError as e:
            raise ExportacionError("PDF", "reportlab no está instalado") from e

        try:
            registros = self._obtener_registros(fecha_inicio, fecha_fin)
            nombre = nombre_archivo or self._generar_nombre("reporte", "pdf")
            ruta = self._directorio / nombre

            doc = SimpleDocTemplate(
                str(ruta),
                pagesize=A4,
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2*cm,
            )

            estilos = getSampleStyleSheet()
            elementos = []

            # Título
            estilo_titulo = ParagraphStyle(
                "Titulo",
                parent=estilos["Title"],
                fontSize=20,
                textColor=colors.HexColor("#1e293b"),
                spaceAfter=12,
            )
            elementos.append(Paragraph(f"{APP_NOMBRE} — Reporte de Salud", estilo_titulo))

            # Subtítulo con fechas
            rango_texto = ""
            if fecha_inicio and fecha_fin:
                rango_texto = f"Del {fecha_inicio} al {fecha_fin}"
            elif fecha_inicio:
                rango_texto = f"Desde el {fecha_inicio}"
            else:
                rango_texto = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"

            elementos.append(Paragraph(rango_texto, estilos["Normal"]))
            elementos.append(Paragraph(f"Total de registros: {len(registros)}", estilos["Normal"]))
            elementos.append(Spacer(1, 0.5*cm))
            elementos.append(HRFlowable(width="100%", color=colors.HexColor("#e2e8f0")))
            elementos.append(Spacer(1, 0.5*cm))

            # Tabla de datos
            encabezados_tabla = [
                "Fecha", "Período", "PA Sys", "PA Dia",
                "FC", "SpO2", "Peso", "IMC", "Pasos",
            ]
            datos_tabla = [encabezados_tabla]

            for r in registros:
                datos_tabla.append([
                    r.fecha.strftime("%d/%m/%Y") if r.fecha else "",
                    PERIODOS.get(r.periodo, r.periodo),
                    f"{r.presion_sistolica}" if r.presion_sistolica else "—",
                    f"{r.presion_diastolica}" if r.presion_diastolica else "—",
                    f"{r.ritmo_cardiaco}" if r.ritmo_cardiaco else "—",
                    f"{r.oxigenacion}%" if r.oxigenacion else "—",
                    f"{r.peso}kg" if r.peso else "—",
                    f"{r.imc}" if r.imc else "—",
                    f"{r.pasos:,}" if r.pasos else "—",
                ])

            tabla = Table(datos_tabla, repeatRows=1)
            tabla.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))

            elementos.append(tabla)

            doc.build(elementos)
            logger.info("PDF exportado: %s (%d registros)", ruta, len(registros))
            return ruta

        except Exception as e:
            raise ExportacionError("PDF", str(e)) from e

    # ──────────────────────────────────────────
    # Helpers
    # ──────────────────────────────────────────

    def _obtener_registros(
        self, fecha_inicio: Optional[date], fecha_fin: Optional[date]
    ):
        """Obtiene los registros del rango especificado."""
        with self._gestor.sesion() as sesion:
            repo = RepositorioRegistro(sesion)
            if fecha_inicio and fecha_fin:
                return repo.obtener_rango(fecha_inicio, fecha_fin)
            return repo.obtener_todos()

    @staticmethod
    def _generar_nombre(prefijo: str, extension: str) -> str:
        """Genera un nombre de archivo con timestamp."""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"healthtrack_{prefijo}_{ts}.{extension}"
